"""Export presentation-ready Block 147 structured data.

Creates:
  * output/block147_building_data.xlsx
  * output/data.html
  * output/data/building_data.json
"""
from __future__ import annotations

import html
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from ..common.paths import PARSED, PROJECT_ROOT


OUTPUT = PROJECT_ROOT / "output"
REPORTS = OUTPUT / "reports"
MANUAL = PROJECT_ROOT / "data" / "manual" / "parcels"

REPLACED_BY_MANUAL = {
    "N-40-42": {"N-40", "N-42"},
    "N-52-54-E2": {"N-52", "N-54", "E-2"},
    "S-41-43-45-E16": {"S-41", "S-43", "S-45", "E-16"},
    "W-34-36-FIRIN": {"W-34", "W-36", "W-38"},
    "W-39-1": {"W-39/1"},
}


CODE_LEGEND = [
    {
        "code": "MO",
        "meaning": "Masonry ordinary / masonry upper-storey notation",
        "model_use": "Usually contributes to masonry class B and upper-storey count.",
        "confidence": "High in this workbook; preserve raw context because Pervititch handwriting varies.",
    },
    {
        "code": "M.",
        "meaning": "Masonry / massive element depending on context",
        "model_use": "Treat with material class and decoded Excel note; not enough alone to place openings.",
        "confidence": "Medium; source rows distinguish M., MB, and MO.",
    },
    {
        "code": "MB",
        "meaning": "Massive brick / masonry body",
        "model_use": "Masonry mass or Firin body provenance.",
        "confidence": "High where decoded row says Massive Brick.",
    },
    {
        "code": "Mg.",
        "meaning": "Magasin: shop/store, sometimes bakery frontage",
        "model_use": "Ground-floor shop/magasin entrance. Does not automatically mean glass vitrine.",
        "confidence": "High.",
    },
    {
        "code": "VF",
        "meaning": "Voute Francaise: French vault",
        "model_use": "Vault/roof provenance; can drive shallow vault/flat roof interpretation.",
        "confidence": "High.",
    },
    {
        "code": "VT",
        "meaning": "Voute Turque: Turkish vault / brick arch vault",
        "model_use": "Vault/ceiling provenance; exterior roof remains conservative unless map shows more.",
        "confidence": "High.",
    },
    {
        "code": "TF",
        "meaning": "Tuiles Francaises: French/Marseille clay roof tiles",
        "model_use": "Rendered as tile_TF / Marseille-style terracotta tile, not sheet metal.",
        "confidence": "High after map recheck.",
    },
    {
        "code": "T.",
        "meaning": "Tole: sheet metal roof/covering",
        "model_use": "Rendered as aged sheet-metal roof when used as roof material.",
        "confidence": "High.",
    },
    {
        "code": "TR / TR.4",
        "meaning": "Turkish clay tile; TR.4 is a specific tile zone/mark on the plan",
        "model_use": "Rendered as tile_TR / terracotta roof, especially Firin complex.",
        "confidence": "High where roof sheet confirms it.",
    },
    {
        "code": "bs. / Bs.",
        "meaning": "Basement / bodrum level",
        "model_use": "Adds basement provenance; not a visible upper storey.",
        "confidence": "High when written as bs.; do not confuse with a plain b suffix.",
    },
    {
        "code": "b",
        "meaning": "Register/map storey suffix; context-dependent",
        "model_use": "Kept as source notation unless a basement is also explicitly marked as bs.",
        "confidence": "Medium; do not overread as basement by itself.",
    },
    {
        "code": "p / P",
        "meaning": "Partial floor / partial storey notation in this register",
        "model_use": "May indicate partial upper floor; only modeled if map geometry supports it.",
        "confidence": "Medium.",
    },
    {
        "code": "E / etage",
        "meaning": "Etage/storey notation, e.g. 3E = three storeys",
        "model_use": "Contributes to storey count.",
        "confidence": "High in rows where decoded as etage.",
    },
    {
        "code": "Vx.",
        "meaning": "Vieux/old condition",
        "model_use": "Condition/material-aging cue; does not add geometry by itself.",
        "confidence": "High.",
    },
    {
        "code": "Vitr. / Vitre / Camli",
        "meaning": "Glazed/vitrine/camli area",
        "model_use": "Only mapped glazed area gets glass; adjacent material zones stay opaque.",
        "confidence": "High where explicitly mapped.",
    },
    {
        "code": "x / x mark",
        "meaning": "Ambiguous map mark: possible tabatiere/skylight or gate",
        "model_use": "Conservative: roof tabatiere/skylight only where supported; not an extra facade window.",
        "confidence": "Medium.",
    },
    {
        "code": "Delta / storey mark",
        "meaning": "Storey count marker in the Excel/register parsing",
        "model_use": "Used as source storey count, then checked against map/manual zones.",
        "confidence": "High as source; geometry still map-checked.",
    },
    {
        "code": "A / B / C",
        "meaning": "Material class: A massive stone, B masonry/plastered brick, C wooden frame",
        "model_use": "Controls wall thickness/material palette and cladding.",
        "confidence": "High when map color agrees.",
    },
]


@dataclass
class Exports:
    workbook: Path
    html_page: Path
    json_data: Path


class DataExportPipeline:
    def run(self) -> Exports:
        parcels = _read_json(PARSED / "parcels.json")
        buildings = _read_json(PARSED / "buildings.json")
        parcel_index = {p["parcel_id"]: p for p in parcels}
        report_tables = {
            "openings": _table_by_id(REPORTS / "opening_audit.md"),
            "roofs": _table_by_id(REPORTS / "roof_visual_audit.md"),
            "geometry": _table_by_id(REPORTS / "geometry_manifest.md"),
        }
        manual_labels = _load_manual_labels()

        model_units = _model_unit_rows(buildings, parcel_index, manual_labels, report_tables)
        source_rows = _source_parcel_rows(parcels)
        manual_rows = _manual_rows(manual_labels)
        qa_rows = _qa_rows(model_units)

        data = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "model_units": model_units,
            "source_parcels": source_rows,
            "manual_overrides": manual_rows,
            "qa_flags": qa_rows,
            "code_legend": _legend_with_occurrences(CODE_LEGEND, source_rows, manual_rows),
        }

        OUTPUT.mkdir(parents=True, exist_ok=True)
        data_dir = OUTPUT / "data"
        data_dir.mkdir(parents=True, exist_ok=True)

        json_path = data_dir / "building_data.json"
        json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

        xlsx_path = OUTPUT / "block147_building_data.xlsx"
        _write_workbook(xlsx_path, data)

        html_path = OUTPUT / "data.html"
        html_path.write_text(_html_page(data), encoding="utf-8")

        print(f"wrote {xlsx_path.relative_to(PROJECT_ROOT)}")
        print(f"wrote {html_path.relative_to(PROJECT_ROOT)}")
        print(f"wrote {json_path.relative_to(PROJECT_ROOT)}")
        return Exports(workbook=xlsx_path, html_page=html_path, json_data=json_path)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def _load_manual_labels() -> list[dict[str, Any]]:
    labels: list[dict[str, Any]] = []
    if not MANUAL.exists():
        return labels
    for path in sorted(MANUAL.glob("*.json")):
        labels.append(_read_json(path))
    return labels


def _model_unit_rows(
    buildings: list[dict[str, Any]],
    parcel_index: dict[str, dict[str, Any]],
    manual_labels: list[dict[str, Any]],
    report_tables: dict[str, dict[str, dict[str, str]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    replaced = set().union(*REPLACED_BY_MANUAL.values())

    for building in buildings:
        pid = building["parcel_id"]
        if pid in replaced:
            continue
        source = parcel_index.get(_base_parcel_id(pid), {})
        rows.append(_row_from_building(building, source, report_tables))

    for label in manual_labels:
        for zone in label.get("zones") or []:
            rows.append(_row_from_manual_zone(label, zone, report_tables))

    rows.extend(_special_asset_rows(report_tables))
    rows.sort(key=lambda r: _sort_key(str(r["model_id"])))
    return rows


def _row_from_building(
    building: dict[str, Any],
    source: dict[str, Any],
    report_tables: dict[str, dict[str, dict[str, str]]],
) -> dict[str, Any]:
    pid = building["parcel_id"]
    roof = building.get("roof") or {}
    openings = report_tables["openings"].get(pid, {})
    roof_report = report_tables["roofs"].get(pid, {})
    geom = report_tables["geometry"].get(pid, {})
    source_roof = source.get("roof") or {}
    source_wall = source.get("wall") or {}
    source_vault = source.get("vault") or {}
    material = source.get("material") or {}
    notes = building.get("notes") or {}
    storeys = [s for s in building.get("storeys") or [] if not s.get("is_basement")]

    return {
        "model_id": pid,
        "source_parcels": _base_parcel_id(pid),
        "source_footprint": ((building.get("provenance") or {}).get("footprint_source_file") or openings.get("footprint")),
        "source_type": "excel+geometry",
        "manual_override": "No",
        "material_class": building.get("material_class") or material.get("class"),
        "material_interpretation": material.get("decoded") or material.get("raw_material_label"),
        "wall_code": (building.get("excel_snapshot") or {}).get("wall_code") or source_wall.get("code"),
        "wall_decoded": source_wall.get("decoded"),
        "vault_code": (building.get("excel_snapshot") or {}).get("vault_code") or source_vault.get("code"),
        "vault_decoded": source_vault.get("decoded"),
        "map_labels": "",
        "source_storeys": ((building.get("excel_snapshot") or {}).get("storeys_raw") or (source.get("storeys") or {}).get("raw")),
        "model_storeys": len(storeys),
        "model_height_m": round(sum(float(s.get("height_m") or 0) for s in storeys), 2),
        "basement": _yes_no(any(s.get("is_basement") for s in building.get("storeys") or []) or source.get("basement")),
        "ground_floor_use": _ground_use(building, source),
        "roof_shape": roof_report.get("shape") or roof.get("shape") or source_roof.get("shape"),
        "roof_material": roof_report.get("source material") or roof.get("material") or source_roof.get("material_code"),
        "roof_pitch_deg": roof_report.get("pitch") or roof.get("pitch_deg"),
        "rendered_roof_materials": roof_report.get("rendered roof materials"),
        "doors": _int(openings.get("doors")),
        "shopfronts": _int(openings.get("shopfronts")),
        "upper_windows": _int(openings.get("upper windows")),
        "strict_street_edges": openings.get("strict street edges"),
        "palette_source": geom.get("palette.source") or ((building.get("facade_palette") or {}).get("source")),
        "roles": geom.get("roles"),
        "source_notes": _join_notes(openings.get("source notes"), roof_report.get("note"), notes),
    }


def _row_from_manual_zone(
    label: dict[str, Any],
    zone: dict[str, Any],
    report_tables: dict[str, dict[str, dict[str, str]]],
) -> dict[str, Any]:
    model_id = f"{label['label']}.{zone['id']}"
    openings = report_tables["openings"].get(model_id, {})
    roof_report = report_tables["roofs"].get(model_id, {})
    geom = report_tables["geometry"].get(model_id, {})
    roof = zone.get("roof") or {}
    source_parcels = ", ".join(label.get("parcel_ids") or [])
    storey_heights = [float(h) for h in zone.get("storey_heights_m") or []]
    material_desc = _material_meaning(zone.get("material_class"), zone.get("map_colour"))

    return {
        "model_id": model_id,
        "source_parcels": source_parcels,
        "source_footprint": label.get("footprint_ref") or openings.get("footprint"),
        "source_type": "manual map-zoned",
        "manual_override": "Yes",
        "material_class": zone.get("material_class"),
        "material_interpretation": material_desc,
        "wall_code": ", ".join(zone.get("map_labels") or []),
        "wall_decoded": zone.get("description"),
        "vault_code": "",
        "vault_decoded": "",
        "map_labels": ", ".join(zone.get("map_labels") or []),
        "source_storeys": zone.get("storeys_above_grade"),
        "model_storeys": len(storey_heights),
        "model_height_m": round(sum(storey_heights), 2),
        "basement": _yes_no(zone.get("has_basement")),
        "ground_floor_use": zone.get("ground_floor_use"),
        "roof_shape": roof_report.get("shape") or roof.get("shape"),
        "roof_material": roof_report.get("source material") or roof.get("material"),
        "roof_pitch_deg": roof_report.get("pitch") or roof.get("pitch_deg"),
        "rendered_roof_materials": roof_report.get("rendered roof materials"),
        "doors": _int(openings.get("doors")),
        "shopfronts": _int(openings.get("shopfronts")),
        "upper_windows": _int(openings.get("upper windows")),
        "strict_street_edges": openings.get("strict street edges"),
        "palette_source": geom.get("palette.source"),
        "roles": geom.get("roles"),
        "source_notes": _join_notes(label.get("map_notes"), roof_report.get("note"), openings.get("source notes")),
    }


def _special_asset_rows(report_tables: dict[str, dict[str, dict[str, str]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pid in ("CHURCH", "COURTYARD-147-GARDEN"):
        geom = report_tables["geometry"].get(pid, {})
        roof = report_tables["roofs"].get(pid, {})
        if not geom and not roof:
            continue
        rows.append({
            "model_id": pid,
            "source_parcels": pid,
            "source_footprint": "special/manual scene asset",
            "source_type": "special asset",
            "manual_override": "Yes",
            "material_class": "",
            "material_interpretation": "Church/special landscape asset" if pid == "CHURCH" else "Courtyard garden / vegetation",
            "wall_code": "",
            "wall_decoded": "",
            "vault_code": "",
            "vault_decoded": "",
            "map_labels": "",
            "source_storeys": "",
            "model_storeys": "",
            "model_height_m": "",
            "basement": "",
            "ground_floor_use": "",
            "roof_shape": roof.get("shape"),
            "roof_material": roof.get("source material"),
            "roof_pitch_deg": roof.get("pitch"),
            "rendered_roof_materials": roof.get("rendered roof materials"),
            "doors": "",
            "shopfronts": "",
            "upper_windows": "",
            "strict_street_edges": "",
            "palette_source": geom.get("palette.source"),
            "roles": geom.get("roles"),
            "source_notes": roof.get("note"),
        })
    return rows


def _source_parcel_rows(parcels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for p in parcels:
        wall = p.get("wall") or {}
        vault = p.get("vault") or {}
        gf = p.get("ground_floor") or {}
        storeys = p.get("storeys") or {}
        roof = p.get("roof") or {}
        material = p.get("material") or {}
        openings = p.get("openings") or {}
        walls = p.get("walls_analysis") or {}
        rows.append({
            "parcel_id": p.get("parcel_id"),
            "parcel_number": p.get("parcel_number"),
            "zone": p.get("zone"),
            "street_facing": p.get("street_facing"),
            "material_class": material.get("class"),
            "material_decoded": material.get("decoded"),
            "map_colour": material.get("map_colour"),
            "wall_code": wall.get("code"),
            "wall_decoded": wall.get("decoded"),
            "vault_code": vault.get("code"),
            "vault_decoded": vault.get("decoded"),
            "ground_floor_use": gf.get("use"),
            "ground_floor_code": gf.get("code"),
            "storeys_raw": storeys.get("raw"),
            "storeys_count": storeys.get("count"),
            "basement": p.get("basement"),
            "condition": p.get("condition"),
            "roof_shape": roof.get("shape"),
            "roof_material_code": roof.get("material_code"),
            "roof_material_decoded": roof.get("material_decoded"),
            "primary_door_face": openings.get("primary_door_face"),
            "secondary_door_face": openings.get("secondary_door_face"),
            "wall_line_feature": walls.get("special_feature"),
            "hatch_meaning": walls.get("hatch_meaning"),
            "bim_notes": _join_notes(p.get("bim_notes"), openings.get("bim_notes"), walls.get("bim_notes")),
        })
    return rows


def _manual_rows(labels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for label in labels:
        for zone in label.get("zones") or []:
            roof = zone.get("roof") or {}
            rows.append({
                "manual_label": label.get("label"),
                "source_parcels": ", ".join(label.get("parcel_ids") or []),
                "zone_id": zone.get("id"),
                "footprint_ref": label.get("footprint_ref"),
                "footprint_mode": label.get("footprint_mode"),
                "clip_ranges": json.dumps(zone.get("clip_ranges") or [], ensure_ascii=False),
                "footprint_fraction": json.dumps(zone.get("footprint_fraction"), ensure_ascii=False),
                "material_class": zone.get("material_class"),
                "map_colour": zone.get("map_colour"),
                "storeys_above_grade": zone.get("storeys_above_grade"),
                "storey_heights_m": ", ".join(str(h) for h in zone.get("storey_heights_m") or []),
                "has_basement": _yes_no(zone.get("has_basement")),
                "ground_floor_use": zone.get("ground_floor_use"),
                "roof_shape": roof.get("shape"),
                "roof_material": roof.get("material"),
                "roof_pitch_deg": roof.get("pitch_deg"),
                "map_labels": ", ".join(zone.get("map_labels") or []),
                "zone_description": zone.get("description"),
                "map_notes": label.get("map_notes"),
            })
    return rows


def _qa_rows(model_units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in model_units:
        flags: list[str] = []
        if not row.get("roof_material") or row.get("roof_material") in ("unknown", "—"):
            flags.append("Roof material unreadable/unknown")
        if row.get("manual_override") == "No" and not row.get("wall_code"):
            flags.append("No wall code in parsed source")
        if row.get("source_type") == "excel+geometry" and row.get("upper_windows") and not row.get("doors"):
            flags.append("Upper windows present but no door count; check opening evidence")
        if row.get("model_id") == "W-39-1.wooden_church_edge_annex":
            flags.append("Intentionally opaque: no glass/openings despite adjacency to Camli area")
        if flags:
            out.append({
                "model_id": row.get("model_id"),
                "flags": "; ".join(flags),
                "notes": row.get("source_notes"),
            })
    return out


def _legend_with_occurrences(
    legend: list[dict[str, str]],
    source_rows: list[dict[str, Any]],
    manual_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    haystack = "\n".join(
        " ".join(str(v or "") for v in row.values())
        for row in [*source_rows, *manual_rows]
    )
    out = []
    for item in legend:
        count = sum(
            len(re.findall(pattern, haystack, flags=re.IGNORECASE))
            for pattern in _legend_patterns(item["code"])
        )
        out.append({**item, "observed_mentions": count})
    return out


def _legend_patterns(code: str) -> list[str]:
    if code == "MO":
        return [r"\b\d*MO\b"]
    if code == "M.":
        return [r"(?<![A-Za-z])M\.(?![A-Za-z])"]
    if code == "MB":
        return [r"\bMB\b", r"\bMassive Brick\b"]
    if code == "Mg.":
        return [r"\bMg\.", r"\bMG\.", r"\bMagasin\b"]
    if code == "VF":
        return [r"\bVF\b", r"\bVoute Francaise\b", r"\bVoûte Française\b"]
    if code == "VT":
        return [r"\bVT\b", r"\bVoute Turque\b", r"\bVoûte Turque\b"]
    if code == "TF":
        return [r"\bTF\b", r"\bTuiles Francaises\b"]
    if code == "T.":
        return [r"(?<![A-Za-z])T\.(?![A-Za-z])", r"\bTole\b", r"\bTôle\b"]
    if code == "TR / TR.4":
        return [r"\bTR(?:\.4)?\b"]
    if code == "bs. / Bs.":
        return [r"\bbs\.?\b"]
    if code == "b":
        return [r"\d+b\b"]
    if code == "p / P":
        return [r"\d+p\b", r"\d+P\b"]
    if code == "E / etage":
        return [r"\d+E\b", r"\betage\b", r"\bétage\b"]
    if code == "Vx.":
        return [r"\bVx\.", r"\bVieux\b"]
    if code == "Vitr. / Vitre / Camli":
        return [r"\bVitr\.?", r"\bVitre\b", r"\bCamli\b", r"\bCamlı\b"]
    if code == "x / x mark":
        return [r"×", r"\bx mark\b"]
    if code == "Delta / storey mark":
        return [r"Δ", r"\bDelta\b"]
    if code == "A / B / C":
        return [r"\bClass A\b", r"\bClass B\b", r"\bClass C\b"]
    return [re.escape(code)]


def _write_workbook(path: Path, data: dict[str, Any]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    summary_rows = [
        {"field": "generated_at", "value": data["generated_at"]},
        {"field": "visible_model_units", "value": len(data["model_units"])},
        {"field": "source_parcels", "value": len(data["source_parcels"])},
        {"field": "manual_override_zones", "value": len(data["manual_overrides"])},
        {"field": "qa_flags", "value": len(data["qa_flags"])},
        {"field": "viewer_data_url", "value": "data.html"},
        {"field": "model_viewer_url", "value": "viewer.html"},
    ]

    _add_sheet(wb, "README", summary_rows)
    _add_sheet(wb, "Code_Legend", data["code_legend"])
    _add_sheet(wb, "Model_Units", data["model_units"])
    _add_sheet(wb, "Source_Parcels", data["source_parcels"])
    _add_sheet(wb, "Manual_Overrides", data["manual_overrides"])
    _add_sheet(wb, "QA_Flags", data["qa_flags"])

    wb.save(path)


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No rows"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_excel_value(row.get(h)) for h in headers])

    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=_safe_table_name(title), ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table_ref
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F2933")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        values = [str(ws.cell(row=r, column=col_idx).value or "") for r in range(1, min(max_row, 80) + 1)]
        width = min(46, max(10, max(len(v) for v in values[:80]) + 2))
        if header in {"source_notes", "bim_notes", "map_notes", "wall_decoded", "zone_description"}:
            width = 58
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _html_page(data: dict[str, Any]) -> str:
    payload = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    generated = html.escape(str(data["generated_at"]))
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Block 147 Building Data</title>
  <style>
    :root {{
      color-scheme: light;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f5f2ec;
      color: #202126;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; }}
    header {{
      position: sticky; top: 0; z-index: 10;
      display: grid; grid-template-columns: 1fr auto; gap: 14px; align-items: center;
      padding: 18px 22px; background: rgba(245, 242, 236, 0.96);
      border-bottom: 1px solid #d8d0c3; backdrop-filter: blur(8px);
    }}
    h1 {{ margin: 0; font-size: 22px; line-height: 1.15; letter-spacing: 0; }}
    .meta {{ margin-top: 4px; color: #686056; font-size: 13px; }}
    nav {{ display: flex; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }}
    a.button, button {{
      border: 1px solid #b8aa97; background: #fffaf2; color: #29241f;
      border-radius: 6px; padding: 8px 11px; font-size: 13px; text-decoration: none; cursor: pointer;
    }}
    button.active {{ background: #2f4152; border-color: #2f4152; color: #fff; }}
    main {{ padding: 18px 22px 40px; }}
    .toolbar {{
      display: grid; grid-template-columns: minmax(220px, 420px) 170px 170px 1fr;
      gap: 10px; align-items: center; margin-bottom: 14px;
    }}
    input, select {{
      width: 100%; border: 1px solid #c8bbab; border-radius: 6px; padding: 9px 10px;
      background: #fff; color: #1f1f1f; font: inherit; font-size: 13px;
    }}
    .count {{ justify-self: end; color: #686056; font-size: 13px; }}
    .table-wrap {{ overflow: auto; border: 1px solid #d7cfc4; background: #fffdf8; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 1400px; }}
    th, td {{ border-bottom: 1px solid #ece5dc; padding: 8px 10px; vertical-align: top; font-size: 12px; }}
    th {{ position: sticky; top: 86px; background: #2f4152; color: #fff; text-align: left; z-index: 2; }}
    tr:nth-child(even) td {{ background: #fbf7ef; }}
    .pill {{ display: inline-block; border: 1px solid #c8bbab; border-radius: 999px; padding: 2px 7px; background: #fff; }}
    .legend-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 10px; }}
    .legend-card {{ border: 1px solid #d7cfc4; background: #fffdf8; padding: 12px; border-radius: 8px; }}
    .legend-code {{ font-weight: 700; font-size: 16px; }}
    .muted {{ color: #746b61; }}
    @media (max-width: 760px) {{
      header {{ grid-template-columns: 1fr; }}
      nav {{ justify-content: start; }}
      .toolbar {{ grid-template-columns: 1fr; }}
      .count {{ justify-self: start; }}
      th {{ top: 154px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Block 147 Building Data</h1>
      <div class="meta">Generated {generated} · source Excel + manual map corrections + current viewer reports</div>
    </div>
    <nav>
      <button class="active" data-view="units">Model Units</button>
      <button data-view="legend">Code Legend</button>
      <button data-view="source">Source Parcels</button>
      <a class="button" href="./block147_building_data.xlsx">Excel</a>
      <a class="button" href="./viewer.html?refresh=data">3D Viewer</a>
    </nav>
  </header>
  <main id="app"></main>
  <script id="payload" type="application/json">{payload}</script>
  <script>
    const data = JSON.parse(document.getElementById('payload').textContent);
    let view = 'units';
    let query = '';
    let material = '';
    let manual = '';
    const app = document.getElementById('app');
    document.querySelectorAll('button[data-view]').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('button[data-view]').forEach(b => b.classList.remove('active'));
        btn.classList.add('active');
        view = btn.dataset.view;
        render();
      }});
    }});

    function render() {{
      if (view === 'legend') return renderLegend();
      const rows = view === 'source' ? data.source_parcels : data.model_units;
      const filtered = rows.filter(row => {{
        const text = Object.values(row).join(' ').toLowerCase();
        const okQuery = !query || text.includes(query.toLowerCase());
        const okMat = !material || String(row.material_class || '').includes(material);
        const okManual = !manual || String(row.manual_override || '') === manual;
        return okQuery && okMat && okManual;
      }});
      app.innerHTML = `
        <div class="toolbar">
          <input id="q" value="${{escapeAttr(query)}}" placeholder="Search parcel, code, roof, note...">
          <select id="mat">
            <option value="">All materials</option>
            <option value="A" ${{material==='A'?'selected':''}}>A stone</option>
            <option value="B" ${{material==='B'?'selected':''}}>B masonry</option>
            <option value="C" ${{material==='C'?'selected':''}}>C wooden</option>
          </select>
          <select id="manual" ${{view === 'source' ? 'disabled' : ''}}>
            <option value="">All source types</option>
            <option value="Yes" ${{manual==='Yes'?'selected':''}}>Manual corrected</option>
            <option value="No" ${{manual==='No'?'selected':''}}>Excel/geometry</option>
          </select>
          <div class="count">${{filtered.length}} rows</div>
        </div>
        ${{table(filtered)}}
      `;
      document.getElementById('q').addEventListener('input', e => {{ query = e.target.value; render(); }});
      document.getElementById('mat').addEventListener('change', e => {{ material = e.target.value; render(); }});
      const manualEl = document.getElementById('manual');
      if (manualEl) manualEl.addEventListener('change', e => {{ manual = e.target.value; render(); }});
    }}

    function renderLegend() {{
      app.innerHTML = `<div class="legend-grid">${{data.code_legend.map(item => `
        <section class="legend-card">
          <div class="legend-code">${{escapeHtml(item.code)}} <span class="muted">(${{item.observed_mentions || 0}} mentions)</span></div>
          <p>${{escapeHtml(item.meaning)}}</p>
          <p><strong>Model:</strong> ${{escapeHtml(item.model_use)}}</p>
          <p class="muted">${{escapeHtml(item.confidence)}}</p>
        </section>
      `).join('')}}</div>`;
    }}

    function table(rows) {{
      if (!rows.length) return '<div class="table-wrap"><table><tbody><tr><td>No rows</td></tr></tbody></table></div>';
      const headers = Object.keys(rows[0]);
      return `<div class="table-wrap"><table><thead><tr>${{headers.map(h => `<th>${{escapeHtml(h)}}</th>`).join('')}}</tr></thead><tbody>
        ${{rows.map(row => `<tr>${{headers.map(h => cell(row[h], h)).join('')}}</tr>`).join('')}}
      </tbody></table></div>`;
    }}

    function cell(value, key) {{
      const v = value == null ? '' : String(value);
      if (['material_class','manual_override','roof_material'].includes(key) && v) return `<td><span class="pill">${{escapeHtml(v)}}</span></td>`;
      return `<td>${{escapeHtml(v).replaceAll('\\n', '<br>')}}</td>`;
    }}
    function escapeHtml(s) {{ return String(s).replace(/[&<>"']/g, ch => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[ch])); }}
    function escapeAttr(s) {{ return escapeHtml(s).replaceAll('"', '&quot;'); }}
    render();
  </script>
</body>
</html>
"""


def _html_page(data: dict[str, Any]) -> str:
    """Generate the interactive data viewer.

    Defined after the legacy renderer so this cleaner layout is the active
    implementation while keeping the export contract unchanged.
    """
    payload = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    generated = html.escape(str(data["generated_at"]))
    page = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Block 147 Building Data</title>
  <style>
    :root {
      color-scheme: light;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f4f5f1;
      color: #202422;
      --panel: #fffdf8;
      --panel-2: #f7f1e8;
      --line: #d7d0c4;
      --line-strong: #a99a88;
      --ink-muted: #635f58;
      --accent: #315f55;
      --accent-2: #6f4731;
      --header: #253330;
    }
    * { box-sizing: border-box; }
    html, body { margin: 0; min-height: 100%; }
    body { background: var(--panel-2); }
    .shell { min-height: 100vh; display: grid; grid-template-rows: auto 1fr; }
    header {
      position: sticky;
      top: 0;
      z-index: 20;
      display: grid;
      grid-template-columns: minmax(260px, 1fr) auto;
      gap: 16px;
      align-items: center;
      padding: 16px 20px 12px;
      background: rgba(244, 245, 241, 0.98);
      border-bottom: 1px solid var(--line);
      backdrop-filter: blur(10px);
    }
    h1 { margin: 0; font-size: 20px; line-height: 1.18; letter-spacing: 0; }
    .meta { margin-top: 4px; color: var(--ink-muted); font-size: 12px; }
    nav { display: flex; gap: 7px; flex-wrap: wrap; justify-content: flex-end; }
    a.button, button {
      min-height: 34px;
      border: 1px solid var(--line-strong);
      background: var(--panel);
      color: #251f1a;
      border-radius: 6px;
      padding: 7px 10px;
      font: inherit;
      font-size: 12px;
      line-height: 1.2;
      text-decoration: none;
      cursor: pointer;
      white-space: nowrap;
    }
    button.active { background: var(--accent); border-color: var(--accent); color: #fff; }
    main {
      display: grid;
      grid-template-rows: auto auto 1fr;
      gap: 12px;
      min-height: 0;
      padding: 14px 20px 18px;
    }
    .stats { display: grid; grid-template-columns: repeat(5, minmax(120px, 1fr)); gap: 8px; }
    .stat {
      min-width: 0;
      border: 1px solid var(--line);
      background: var(--panel);
      border-radius: 8px;
      padding: 9px 11px;
    }
    .stat-value { font-size: 18px; font-weight: 700; color: var(--header); }
    .stat-label { margin-top: 2px; color: var(--ink-muted); font-size: 11px; }
    .toolbar {
      display: grid;
      grid-template-columns: minmax(260px, 1fr) 160px 170px 160px auto;
      gap: 9px;
      align-items: center;
    }
    input, select {
      width: 100%;
      min-height: 36px;
      border: 1px solid #c8bbab;
      border-radius: 6px;
      padding: 8px 10px;
      background: #fff;
      color: #1f1f1f;
      font: inherit;
      font-size: 12px;
    }
    .count { justify-self: end; color: var(--ink-muted); font-size: 12px; white-space: nowrap; }
    .table-wrap {
      min-height: 320px;
      max-height: calc(100vh - 218px);
      overflow: auto;
      border: 1px solid var(--line);
      background: var(--panel);
      border-radius: 8px;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.75);
    }
    table { width: 100%; min-width: 1680px; border-collapse: separate; border-spacing: 0; table-layout: fixed; }
    th, td {
      border-bottom: 1px solid #ece5dc;
      border-right: 1px solid #efe8df;
      padding: 8px 10px;
      vertical-align: top;
      font-size: 12px;
      line-height: 1.35;
    }
    th {
      position: sticky;
      top: 0;
      z-index: 3;
      background: var(--header);
      color: #fff;
      text-align: left;
      font-size: 11px;
      font-weight: 650;
      box-shadow: 0 1px 0 rgba(0,0,0,0.18);
    }
    tr:nth-child(even) td { background: #fbf8f2; }
    tr:hover td { background: #f2f5ee; }
    .cell { max-height: 118px; overflow: auto; overflow-wrap: anywhere; }
    .cell.long { min-width: 260px; max-height: 148px; }
    .cell.id { font-weight: 650; color: #1f2c29; }
    .pill {
      display: inline-block;
      border: 1px solid #c4b6a2;
      border-radius: 999px;
      padding: 2px 7px;
      background: #fff;
      color: #332b24;
      white-space: nowrap;
    }
    .pill.Yes { background: #e7f2ee; border-color: #95b7ab; color: #20473f; }
    .pill.No { background: #f5efe7; border-color: #d0b99f; color: #63432c; }
    .pill.A { background: #eee8dc; }
    .pill.B { background: #f5e4da; }
    .pill.C { background: #edf0db; }
    .legend-grid {
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(260px, 1fr));
      gap: 10px;
      overflow: auto;
      max-height: calc(100vh - 170px);
    }
    .legend-card {
      border: 1px solid var(--line);
      background: var(--panel);
      padding: 12px;
      border-radius: 8px;
    }
    .legend-code { font-weight: 700; font-size: 15px; color: var(--accent-2); }
    .legend-card p { margin: 8px 0 0; font-size: 12px; line-height: 1.4; }
    .empty { padding: 24px; color: var(--ink-muted); font-size: 13px; }
    .muted { color: var(--ink-muted); }
    @media (max-width: 980px) {
      header { grid-template-columns: 1fr; }
      nav { justify-content: start; }
      .stats { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .toolbar { grid-template-columns: 1fr 1fr; }
      .count { justify-self: start; }
      .table-wrap { max-height: calc(100vh - 330px); }
    }
    @media (max-width: 620px) {
      header { padding: 14px 14px 10px; }
      main { padding: 12px 14px 16px; }
      .stats, .toolbar { grid-template-columns: 1fr; }
      nav { gap: 6px; }
      .table-wrap { max-height: calc(100vh - 430px); }
    }
  </style>
</head>
<body>
  <div class="shell">
    <header>
      <div>
        <h1>Block 147 Building Data</h1>
        <div class="meta">Generated __GENERATED__ · Excel + manual map corrections + current viewer reports</div>
      </div>
      <nav>
        <button class="active" data-view="units">Model Units</button>
        <button data-view="source">Source Parcels</button>
        <button data-view="manual">Manual Zones</button>
        <button data-view="qa">QA Flags</button>
        <button data-view="legend">Code Legend</button>
        <a class="button" href="./block147_building_data.xlsx">Excel</a>
        <a class="button" href="./viewer.html?refresh=data">3D Viewer</a>
      </nav>
    </header>
    <main id="app"></main>
  </div>
  <script id="payload" type="application/json">__PAYLOAD__</script>
  <script>
    const data = JSON.parse(document.getElementById('payload').textContent);
    let view = 'units';
    let query = '';
    let material = '';
    let manual = '';
    let roof = '';
    const app = document.getElementById('app');
    const viewMap = {
      units: { rows: () => data.model_units || [] },
      source: { rows: () => data.source_parcels || [] },
      manual: { rows: () => data.manual_overrides || [] },
      qa: { rows: () => data.qa_flags || [] }
    };

    document.querySelectorAll('button[data-view]').forEach(btn => {
      btn.addEventListener('click', () => {
        document.querySelectorAll('button[data-view]').forEach(b => b.classList.remove('active'));
        btn.classList.add('active');
        view = btn.dataset.view;
        render();
      });
    });

    function render() {
      if (view === 'legend') return renderLegend();
      const rows = viewMap[view].rows();
      const filtered = rows.filter(row => {
        const text = Object.values(row).join(' ').toLowerCase();
        const okQuery = !query || text.includes(query.toLowerCase());
        const okMat = !material || String(row.material_class || '').includes(material);
        const okManual = !manual || String(row.manual_override || '') === manual;
        const roofText = `${row.roof_shape || ''} ${row.roof_material || ''} ${row.rendered_roof_materials || ''}`.toLowerCase();
        const okRoof = !roof || roofText.includes(roof.toLowerCase());
        return okQuery && okMat && okManual && okRoof;
      });
      app.innerHTML = `
        ${summary()}
        <div class="toolbar">
          <input id="q" value="${escapeAttr(query)}" placeholder="Search parcel, code, roof, note...">
          <select id="mat">
            <option value="">All materials</option>
            <option value="A" ${material==='A'?'selected':''}>A stone</option>
            <option value="B" ${material==='B'?'selected':''}>B masonry</option>
            <option value="C" ${material==='C'?'selected':''}>C wooden</option>
          </select>
          <select id="manual" ${view === 'source' || view === 'manual' || view === 'qa' ? 'disabled' : ''}>
            <option value="">All source types</option>
            <option value="Yes" ${manual==='Yes'?'selected':''}>Manual corrected</option>
            <option value="No" ${manual==='No'?'selected':''}>Excel/geometry</option>
          </select>
          <select id="roof">
            <option value="">All roofs</option>
            <option value="gable" ${roof==='gable'?'selected':''}>Gable</option>
            <option value="hip" ${roof==='hip'?'selected':''}>Hip</option>
            <option value="vault" ${roof==='vault'?'selected':''}>Vault</option>
            <option value="tile" ${roof==='tile'?'selected':''}>Tile</option>
            <option value="sheet" ${roof==='sheet'?'selected':''}>Sheet metal</option>
          </select>
          <div class="count">${filtered.length} rows</div>
        </div>
        ${table(filtered)}
      `;
      document.getElementById('q').addEventListener('input', e => { query = e.target.value; render(); });
      document.getElementById('mat').addEventListener('change', e => { material = e.target.value; render(); });
      document.getElementById('manual').addEventListener('change', e => { manual = e.target.value; render(); });
      document.getElementById('roof').addEventListener('change', e => { roof = e.target.value; render(); });
    }

    function summary() {
      const stats = [
        ['Model units', (data.model_units || []).length],
        ['Source parcels', (data.source_parcels || []).length],
        ['Manual zones', (data.manual_overrides || []).length],
        ['QA flags', (data.qa_flags || []).length],
        ['Code entries', (data.code_legend || []).length],
      ];
      return `<section class="stats">${stats.map(([label, value]) => `
        <div class="stat">
          <div class="stat-value">${escapeHtml(value)}</div>
          <div class="stat-label">${escapeHtml(label)}</div>
        </div>
      `).join('')}</section>`;
    }

    function renderLegend() {
      app.innerHTML = `<div class="legend-grid">${(data.code_legend || []).map(item => `
        <section class="legend-card">
          <div class="legend-code">${escapeHtml(item.code)} <span class="muted">(${item.observed_mentions || 0} mentions)</span></div>
          <p>${escapeHtml(item.meaning)}</p>
          <p><strong>Model:</strong> ${escapeHtml(item.model_use)}</p>
          <p class="muted">${escapeHtml(item.confidence)}</p>
        </section>
      `).join('')}</div>`;
    }

    function table(rows) {
      if (!rows.length) return '<div class="table-wrap"><div class="empty">No rows match the current filters.</div></div>';
      const headers = Object.keys(rows[0]);
      return `<div class="table-wrap"><table><thead><tr>${headers.map(h => `<th>${escapeHtml(h)}</th>`).join('')}</tr></thead><tbody>
        ${rows.map(row => `<tr>${headers.map(h => cell(row[h], h)).join('')}</tr>`).join('')}
      </tbody></table></div>`;
    }

    function cell(value, key) {
      const v = value == null ? '' : String(value);
      const display = escapeHtml(v).replaceAll('\\n', '<br>');
      const longKeys = ['source_notes', 'map_notes', 'zone_description', 'bim_notes', 'wall_decoded', 'vault_decoded', 'roles', 'rendered_roof_materials'];
      const idKeys = ['model_id', 'parcel_id', 'manual_label', 'source_parcels', 'source_footprint'];
      if (['material_class','manual_override','roof_material'].includes(key) && v) {
        return `<td><span class="pill ${escapeAttr(v)}">${display}</span></td>`;
      }
      const klass = `${longKeys.includes(key) ? 'long ' : ''}${idKeys.includes(key) ? 'id ' : ''}`.trim();
      return `<td><div class="cell ${klass}">${display}</div></td>`;
    }
    function escapeHtml(s) { return String(s).replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch])); }
    function escapeAttr(s) { return escapeHtml(s).replaceAll('"', '&quot;'); }
    render();
  </script>
</body>
</html>
"""
    return page.replace("__GENERATED__", generated).replace("__PAYLOAD__", payload)


def _html_page(data: dict[str, Any]) -> str:
    """Generate a readable master-detail data viewer, not a 30-column squeeze."""
    payload = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    generated = html.escape(str(data["generated_at"]))
    page = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Block 147 Building Data</title>
  <style>
    :root {
      color-scheme: light;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      --bg: #f3efe7;
      --panel: #fffdf8;
      --panel-soft: #f8f5ef;
      --ink: #202422;
      --muted: #625d55;
      --line: #d8cec0;
      --line-strong: #a99480;
      --green: #315f55;
      --brick: #7a4d34;
      --blue: #2f586b;
    }
    * { box-sizing: border-box; }
    html, body { margin: 0; min-height: 100%; background: var(--bg); color: var(--ink); }
    body { overflow: hidden; }
    .shell { height: 100vh; display: grid; grid-template-rows: auto 1fr; }
    header {
      display: grid;
      grid-template-columns: minmax(280px, 1fr) auto;
      gap: 16px;
      align-items: center;
      padding: 16px 22px 14px;
      background: rgba(255,253,248,0.98);
      border-bottom: 1px solid var(--line);
    }
    h1 { margin: 0; font-size: 21px; line-height: 1.16; letter-spacing: 0; }
    .meta { margin-top: 4px; color: var(--muted); font-size: 12px; }
    nav { display: flex; gap: 7px; flex-wrap: wrap; justify-content: flex-end; }
    button, .button {
      min-height: 36px;
      border: 1px solid var(--line-strong);
      border-radius: 6px;
      padding: 8px 11px;
      background: #fffaf3;
      color: #2b241e;
      font: inherit;
      font-size: 12px;
      line-height: 1.2;
      text-decoration: none;
      cursor: pointer;
      white-space: nowrap;
    }
    button.active { background: var(--green); color: #fff; border-color: var(--green); }
    main {
      min-height: 0;
      display: grid;
      grid-template-rows: auto auto 1fr;
      gap: 12px;
      padding: 14px 22px 18px;
    }
    .stats { display: grid; grid-template-columns: repeat(5, minmax(130px, 1fr)); gap: 10px; }
    .stat {
      min-width: 0;
      padding: 10px 12px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .stat-value { font-size: 22px; font-weight: 750; color: #263530; }
    .stat-label { margin-top: 2px; color: var(--muted); font-size: 12px; }
    .toolbar {
      display: grid;
      grid-template-columns: minmax(260px, 1fr) 160px 170px 160px auto;
      gap: 10px;
      align-items: center;
    }
    input, select {
      width: 100%;
      min-height: 40px;
      border: 1px solid #c7b9a6;
      border-radius: 7px;
      padding: 9px 11px;
      background: #fff;
      color: #1f1f1f;
      font: inherit;
      font-size: 13px;
    }
    .count { justify-self: end; color: var(--muted); font-size: 13px; white-space: nowrap; }
    .workspace {
      min-height: 0;
      display: grid;
      grid-template-columns: minmax(520px, 1fr) minmax(360px, 0.72fr);
      gap: 12px;
    }
    .records, .detail {
      min-height: 0;
      overflow: auto;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 9px;
    }
    .records { padding: 8px; }
	    .record {
	      width: 100%;
	      display: block;
	      margin: 0 0 8px;
	      padding: 12px 14px;
	      border: 1px solid #e1d8cc;
	      border-radius: 8px;
	      background: #fffbf4;
	      text-align: left;
	      color: inherit;
	      cursor: pointer;
	      overflow: hidden;
	      white-space: normal;
	    }
	    .record:hover { border-color: #bca891; background: #fbf7ef; }
	    .record.selected { border-color: var(--green); box-shadow: inset 4px 0 0 var(--green); }
	    .record-main {
	      display: grid;
	      grid-template-columns: minmax(0, 1fr) auto;
	      gap: 12px;
	      align-items: start;
	    }
	    .record-title { display: block; font-weight: 760; font-size: 15px; overflow-wrap: anywhere; }
	    .record-subtitle {
	      display: block;
	      margin-top: 3px;
	      color: var(--muted);
	      font-size: 12px;
	      line-height: 1.3;
	      overflow-wrap: anywhere;
	    }
	    .record-fields {
	      display: grid;
	      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
	      gap: 11px 14px;
	      margin-top: 12px;
	    }
	    .field { min-width: 0; }
	    .field-label {
	      display: block;
	      margin-bottom: 3px;
      color: var(--muted);
      font-size: 10px;
      font-weight: 700;
      letter-spacing: .02em;
      text-transform: uppercase;
    }
	    .field-value {
	      display: block;
	      min-width: 0;
	      font-size: 13px;
	      line-height: 1.32;
	      overflow-wrap: anywhere;
	      white-space: normal;
	    }
	    .record-fields .field-value {
	      display: -webkit-box;
	      -webkit-line-clamp: 2;
	      -webkit-box-orient: vertical;
	      overflow: hidden;
	    }
	    .record-notes {
	      min-width: 0;
	      margin-top: 12px;
	      padding-top: 10px;
	      border-top: 1px solid #eadfd2;
	    }
	    .summary-note {
	      display: -webkit-box;
	      -webkit-line-clamp: 2;
	      -webkit-box-orient: vertical;
	      overflow: hidden;
	    }
    .pill {
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      border: 1px solid #c6b7a5;
      border-radius: 999px;
      padding: 2px 8px;
	      background: #fff;
	      color: #332b24;
	      font-size: 12px;
	      white-space: normal;
	      overflow-wrap: anywhere;
	      max-width: 100%;
	    }
    .pill.yes { background: #e7f2ee; border-color: #91b5a8; color: #21483f; }
    .pill.no { background: #f7efe5; border-color: #d0b99f; color: #684731; }
    .pill.a { background: #eee8dc; }
    .pill.b { background: #f5e4da; }
    .pill.c { background: #edf0db; }
    .detail {
      display: grid;
      grid-template-rows: auto 1fr;
      overflow: hidden;
    }
    .detail-head {
      padding: 15px 16px;
      border-bottom: 1px solid var(--line);
      background: #fbf8f1;
    }
    .detail-title { margin: 0; font-size: 18px; line-height: 1.25; overflow-wrap: anywhere; }
    .detail-subtitle { margin-top: 4px; color: var(--muted); font-size: 12px; line-height: 1.35; }
    .kv {
      overflow: auto;
      padding: 10px 12px 14px;
      display: grid;
      grid-template-columns: minmax(130px, 0.35fr) minmax(220px, 1fr);
      gap: 0;
    }
    .kv-key, .kv-value {
      border-bottom: 1px solid #eee5da;
      padding: 9px 8px;
      font-size: 12px;
      line-height: 1.38;
      overflow-wrap: anywhere;
    }
    .kv-key {
      color: var(--muted);
      font-weight: 720;
      background: #fbf8f2;
    }
    .kv-value { white-space: pre-wrap; }
    .legend-grid {
      min-height: 0;
      overflow: auto;
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
      gap: 10px;
    }
    .legend-card {
      border: 1px solid var(--line);
      background: var(--panel);
      border-radius: 8px;
      padding: 13px;
    }
    .legend-code { font-weight: 800; color: var(--brick); }
    .legend-card p { margin: 8px 0 0; font-size: 13px; line-height: 1.4; }
    .empty {
      padding: 24px;
      color: var(--muted);
      font-size: 13px;
    }
	    @media (max-width: 1200px) {
	      body { overflow: auto; }
	      .shell { min-height: 100vh; height: auto; }
	      .workspace { grid-template-columns: 1fr; }
	      .detail { min-height: 520px; }
	      .record-fields { grid-template-columns: repeat(2, minmax(0, 1fr)); }
	    }
	    @media (max-width: 760px) {
	      header { grid-template-columns: 1fr; padding: 14px; }
	      nav { justify-content: start; }
	      main { padding: 12px 14px 16px; }
	      .stats, .toolbar, .record-main, .record-fields, .kv { grid-template-columns: 1fr; }
	      .count { justify-self: start; }
	      .detail { min-height: 620px; }
	    }
  </style>
</head>
<body>
  <div class="shell">
    <header>
      <div>
        <h1>Block 147 Building Data</h1>
        <div class="meta">Generated __GENERATED__ · readable data viewer + Excel export</div>
      </div>
      <nav>
        <button class="active" data-view="units">Model Units</button>
        <button data-view="source">Source Parcels</button>
        <button data-view="manual">Manual Zones</button>
        <button data-view="qa">QA Flags</button>
        <button data-view="legend">Code Legend</button>
        <a class="button" href="./block147_building_data.xlsx">Excel</a>
        <a class="button" href="./viewer.html?refresh=data">3D Viewer</a>
      </nav>
    </header>
    <main id="app"></main>
  </div>
  <script id="payload" type="application/json">__PAYLOAD__</script>
  <script>
    const data = JSON.parse(document.getElementById('payload').textContent);
    let view = 'units';
    let query = '';
    let material = '';
    let manual = '';
    let roof = '';
    let selectedIndex = 0;
    const app = document.getElementById('app');

    const viewConfig = {
      units: {
        rows: () => data.model_units || [],
        title: row => row.model_id || row.parcel_id || 'Model unit',
        subtitle: row => [row.source_parcels, row.source_type].filter(Boolean).join(' · '),
        fields: [
          ['material_class', 'Material'],
          ['model_storeys', 'Storeys'],
          ['roof_shape', 'Roof shape'],
          ['roof_material', 'Roof material'],
          ['openings', 'Openings'],
          ['source_notes', 'Notes']
        ]
      },
      source: {
        rows: () => data.source_parcels || [],
        title: row => row.parcel_id || 'Source parcel',
        subtitle: row => [row.zone, row.street_facing].filter(Boolean).join(' · '),
        fields: [
          ['material_class', 'Material'],
          ['storeys_corrected', 'Storeys'],
          ['roof_shape', 'Roof shape'],
          ['roof_material_decoded', 'Roof material'],
          ['primary_door_face', 'Primary door'],
          ['bim_notes', 'Notes']
        ]
      },
      manual: {
        rows: () => data.manual_overrides || [],
        title: row => `${row.manual_label || ''}.${row.zone_id || ''}`,
        subtitle: row => row.source_parcels || '',
        fields: [
          ['material_class', 'Material'],
          ['storeys_above_grade', 'Storeys'],
          ['roof_shape', 'Roof shape'],
          ['roof_material', 'Roof material'],
          ['map_labels', 'Map labels'],
          ['zone_description', 'Description']
        ]
      },
      qa: {
        rows: () => data.qa_flags || [],
        title: row => row.model_id || 'QA row',
        subtitle: row => row.flags || '',
        fields: [
          ['flags', 'Flags'],
          ['notes', 'Notes']
        ]
      }
    };

    document.querySelectorAll('button[data-view]').forEach(btn => {
      btn.addEventListener('click', () => {
        document.querySelectorAll('button[data-view]').forEach(b => b.classList.remove('active'));
        btn.classList.add('active');
        view = btn.dataset.view;
        selectedIndex = 0;
        render();
      });
    });

    function rowsForView() {
      if (view === 'legend') return [];
      return viewConfig[view].rows();
    }

    function filteredRows() {
      return rowsForView().filter(row => {
        const text = Object.values(row).join(' ').toLowerCase();
        const roofText = `${row.roof_shape || ''} ${row.roof_material || ''} ${row.roof_material_decoded || ''} ${row.rendered_roof_materials || ''}`.toLowerCase();
        const okQuery = !query || text.includes(query.toLowerCase());
        const okMat = !material || String(row.material_class || '').includes(material);
        const okManual = !manual || String(row.manual_override || '') === manual;
        const okRoof = !roof || roofText.includes(roof.toLowerCase());
        return okQuery && okMat && okManual && okRoof;
      });
    }

    function render() {
      if (view === 'legend') return renderLegend();
      const rows = filteredRows();
      if (selectedIndex >= rows.length) selectedIndex = 0;
      const selected = rows[selectedIndex] || null;
      app.innerHTML = `
        ${summary()}
        ${toolbar(rows.length)}
        <section class="workspace">
          <div class="records">${rows.length ? rows.map((row, i) => record(row, i)).join('') : '<div class="empty">No rows match the current filters.</div>'}</div>
          ${detail(selected)}
        </section>
      `;
      bindControls();
    }

    function summary() {
      const stats = [
        ['Model units', (data.model_units || []).length],
        ['Source parcels', (data.source_parcels || []).length],
        ['Manual zones', (data.manual_overrides || []).length],
        ['QA flags', (data.qa_flags || []).length],
        ['Code entries', (data.code_legend || []).length],
      ];
      return `<section class="stats">${stats.map(([label, value]) => `
        <div class="stat">
          <div class="stat-value">${escapeHtml(value)}</div>
          <div class="stat-label">${escapeHtml(label)}</div>
        </div>
      `).join('')}</section>`;
    }

    function toolbar(count) {
      return `<div class="toolbar">
        <input id="q" value="${escapeAttr(query)}" placeholder="Search parcel, code, roof, note...">
        <select id="mat">
          <option value="">All materials</option>
          <option value="A" ${material==='A'?'selected':''}>A stone</option>
          <option value="B" ${material==='B'?'selected':''}>B masonry</option>
          <option value="C" ${material==='C'?'selected':''}>C wooden</option>
        </select>
        <select id="manual" ${view !== 'units' ? 'disabled' : ''}>
          <option value="">All source types</option>
          <option value="Yes" ${manual==='Yes'?'selected':''}>Manual corrected</option>
          <option value="No" ${manual==='No'?'selected':''}>Excel/geometry</option>
        </select>
        <select id="roof">
          <option value="">All roofs</option>
          <option value="gable" ${roof==='gable'?'selected':''}>Gable</option>
          <option value="hip" ${roof==='hip'?'selected':''}>Hip</option>
          <option value="vault" ${roof==='vault'?'selected':''}>Vault</option>
          <option value="tile" ${roof==='tile'?'selected':''}>Tile</option>
          <option value="sheet" ${roof==='sheet'?'selected':''}>Sheet metal</option>
        </select>
        <div class="count">${count} rows</div>
      </div>`;
    }

	    function record(row, i) {
	      const cfg = viewConfig[view];
	      const fields = cfg.fields.slice(0, -1);
	      const note = cfg.fields[cfg.fields.length - 1];
	      return `<button class="record ${i === selectedIndex ? 'selected' : ''}" data-row="${i}">
	        <div class="record-main">
	          <div>
	            <span class="field-label">Record</span>
	            <span class="record-title">${escapeHtml(cfg.title(row))}</span>
	            <span class="record-subtitle">${escapeHtml(cfg.subtitle(row))}</span>
	          </div>
	          ${pillHtml(view === 'units' ? value(row, 'manual_override') : view)}
	        </div>
	        <div class="record-fields">${fields.map(([key, label]) => fieldBlock(row, key, label)).join('')}</div>
	        <div class="record-notes">
	          <span class="field-label">${escapeHtml(note[1])}</span>
	          <span class="field-value summary-note">${escapeHtml(value(row, note[0]))}</span>
	        </div>
	      </button>`;
    }

    function fieldBlock(row, key, label) {
      let v = key === 'openings'
        ? `Doors ${value(row, 'doors') || 0} · Shopfronts ${value(row, 'shopfronts') || 0} · Windows ${value(row, 'upper_windows') || 0}`
        : value(row, key);
	      const pill = ['material_class', 'manual_override', 'roof_material'].includes(key);
	      return `<div class="field">
	        <span class="field-label">${escapeHtml(label)}</span>
	        <span class="field-value">${pill ? pillHtml(v) : escapeHtml(v)}</span>
	      </div>`;
    }

    function detail(row) {
      if (!row) return '<aside class="detail"><div class="empty">Select a row to inspect all fields.</div></aside>';
      const cfg = viewConfig[view];
      const entries = Object.entries(row);
      return `<aside class="detail">
        <div class="detail-head">
          <h2 class="detail-title">${escapeHtml(cfg.title(row))}</h2>
          <div class="detail-subtitle">${escapeHtml(cfg.subtitle(row))}</div>
        </div>
        <div class="kv">${entries.map(([key, val]) => `
          <div class="kv-key">${escapeHtml(key)}</div>
          <div class="kv-value">${formatValue(key, val)}</div>
        `).join('')}</div>
      </aside>`;
    }

    function renderLegend() {
      app.innerHTML = `<div class="legend-grid">${(data.code_legend || []).map(item => `
        <section class="legend-card">
          <div class="legend-code">${escapeHtml(item.code)} <span class="muted">(${item.observed_mentions || 0} mentions)</span></div>
          <p>${escapeHtml(item.meaning)}</p>
          <p><strong>Model:</strong> ${escapeHtml(item.model_use)}</p>
          <p class="muted">${escapeHtml(item.confidence)}</p>
        </section>
      `).join('')}</div>`;
    }

	    function bindControls() {
	      document.getElementById('q').addEventListener('input', e => { query = e.target.value; selectedIndex = 0; render(); });
	      document.getElementById('mat').addEventListener('change', e => { material = e.target.value; selectedIndex = 0; render(); });
	      document.getElementById('manual').addEventListener('change', e => { manual = e.target.value; selectedIndex = 0; render(); });
	      document.getElementById('roof').addEventListener('change', e => { roof = e.target.value; selectedIndex = 0; render(); });
	      document.querySelectorAll('[data-row]').forEach(btn => {
	        btn.addEventListener('click', () => {
	          selectedIndex = Number(btn.dataset.row || 0);
	          updateSelectionOnly();
	        });
	      });
	    }

	    function updateSelectionOnly() {
	      const rows = filteredRows();
	      document.querySelectorAll('[data-row]').forEach(btn => {
	        btn.classList.toggle('selected', Number(btn.dataset.row || 0) === selectedIndex);
	      });
	      const detailEl = document.querySelector('.detail');
	      if (detailEl) {
	        detailEl.outerHTML = detail(rows[selectedIndex] || null);
	      }
	    }

	    function value(row, key) {
      const v = row[key];
      return v == null || v === '' ? '—' : String(v);
    }
    function formatValue(key, val) {
      const v = val == null || val === '' ? '—' : String(val);
      if (['material_class', 'manual_override', 'roof_material'].includes(key) && v !== '—') return pillHtml(v);
      return escapeHtml(v);
    }
    function pillHtml(v) {
      const cls = String(v).toLowerCase().replace(/[^a-z0-9_-]+/g, '');
      return `<span class="pill ${cls}">${escapeHtml(v)}</span>`;
    }
    function escapeHtml(s) { return String(s).replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch])); }
    function escapeAttr(s) { return escapeHtml(s).replaceAll('"', '&quot;'); }
    render();
  </script>
</body>
</html>
"""
    return page.replace("__GENERATED__", generated).replace("__PAYLOAD__", payload)


def _table_by_id(path: Path) -> dict[str, dict[str, str]]:
    rows = _markdown_table(path)
    return {row.get("parcel_id", ""): row for row in rows if row.get("parcel_id")}


def _markdown_table(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    lines = path.read_text().splitlines()
    header: list[str] | None = None
    rows: list[dict[str, str]] = []
    for line in lines:
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if not cells:
            continue
        if all(set(c) <= {"-", ":"} for c in cells):
            continue
        if header is None:
            header = [_normalize_header(c) for c in cells]
            continue
        if len(cells) != len(header):
            continue
        row = {h: _clean_cell(c) for h, c in zip(header, cells)}
        rows.append(row)
    return rows


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def _clean_cell(value: str) -> str:
    return html.unescape(value.replace("<br>", "\n").replace("**", "")).strip()


def _safe_table_name(title: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", title)
    return (name[:240] or "Table") + "_tbl"


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _base_parcel_id(pid: str) -> str:
    return pid.split("#", 1)[0]


def _sort_key(pid: str) -> tuple[str, int, str]:
    match = re.match(r"([A-Z]+)-(\d+)", pid)
    if not match:
        return (pid, 999, "")
    return (match.group(1), int(match.group(2)), pid)


def _ground_use(building: dict[str, Any], source: dict[str, Any]) -> str:
    for storey in building.get("storeys") or []:
        if storey.get("level") == 0 and storey.get("use"):
            return storey["use"]
    return ((source.get("ground_floor") or {}).get("use") or "")


def _material_meaning(cls: str | None, colour: str | None) -> str:
    cls = (cls or "").upper()
    if cls == "A":
        return "Massive stone / high masonry"
    if cls == "B":
        return "Masonry / plastered brick"
    if cls.startswith("C+"):
        return "Wooden with explicit glazed area"
    if cls == "C":
        return "Wooden frame / yellow map class"
    return colour or ""


def _int(value: Any) -> int | str:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return ""


def _yes_no(value: Any) -> str:
    if value in (None, "", "—"):
        return "No"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    s = str(value).strip().lower()
    if s in ("none", "no", "false", "0", "—"):
        return "No"
    return "Yes"


def _join_notes(*parts: Any) -> str:
    out: list[str] = []
    for part in parts:
        if not part:
            continue
        if isinstance(part, dict):
            text = "; ".join(f"{k}: {v}" for k, v in part.items())
        else:
            text = str(part)
        text = text.replace("<br>", "\n").strip()
        if text and text not in out:
            out.append(text)
    return "\n".join(out)


def main() -> None:
    DataExportPipeline().run()


if __name__ == "__main__":
    main()
